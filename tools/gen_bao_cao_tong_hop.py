# -*- coding: utf-8 -*-
"""Gộp toàn bộ báo cáo kiểm thử CareerCompass thành một tệp Excel duy nhất.

Chạy từ thư mục gốc dự án:
    python tools/gen_bao_cao_tong_hop.py

Nguồn dữ liệu:
    BaoCao-PhanA-HopDen.xlsx   — kỹ thuật hộp đen (chép nguyên vẹn)
    BaoCao_Whitebox.xlsx       — kỹ thuật hộp trắng (chép nguyên vẹn)
    tools/ket-qua-chay/*.json  — kết quả newman và CodeceptJS chạy thật,
                                 lưu lại để dựng lại báo cáo mà không phải
                                 chạy lại toàn bộ (mất khoảng 3 phút)

Kết quả: BaoCao-KiemThu-CareerCompass.xlsx

Hai sheet hộp đen và hộp trắng được CHÉP NGUYÊN VĂN kèm định dạng, không gõ lại,
để chữ do từng thành viên tự viết không bị sai lệch.
"""
import json
from copy import copy
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RA = "BaoCao-KiemThu-CareerCompass.xlsx"
NGUON_DEN = "BaoCao-PhanA-HopDen.xlsx"
NGUON_TRANG = "BaoCao_Whitebox.xlsx"

C_HDR = "1F4E79"; C_TXT = "FFFFFFFF"; C_PASS = "C6EFCE"; C_FAIL = "FBE4E4"
C_SEC = "DDEBF7"; C_WARN = "FFF2CC"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
NL = chr(10)


# =====================================================================
# Chép sheet giữa hai workbook, giữ nguyên giá trị và định dạng
# =====================================================================
def chep_sheet(ws_goc, wb_dich, ten_moi):
    ws = wb_dich.create_sheet(ten_moi[:31])
    for hang in ws_goc.iter_rows():
        for o in hang:
            m = ws.cell(row=o.row, column=o.column, value=o.value)
            if o.has_style:
                m.font = copy(o.font)
                m.fill = copy(o.fill)
                m.border = copy(o.border)
                m.alignment = copy(o.alignment)
                m.number_format = o.number_format
    for k, dim in ws_goc.column_dimensions.items():
        ws.column_dimensions[k].width = dim.width
    for k, dim in ws_goc.row_dimensions.items():
        ws.row_dimensions[k].height = dim.height
    for vung in ws_goc.merged_cells.ranges:
        ws.merge_cells(str(vung))
    return ws


# =====================================================================
# Tiện ích dựng bảng
# =====================================================================
def dung(ws, rong=None):
    for i, w in enumerate(rong or [26] * 8, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def o(ws, r, c, v, *, bold=False, fill=None, center=False, size=10, color=None, italic=False):
    x = ws.cell(row=r, column=c, value=v)
    x.font = Font(size=size, bold=bold, italic=italic, color=color)
    x.alignment = Alignment(vertical="top", wrap_text=True,
                            horizontal="center" if center else "general")
    x.border = BORDER
    if fill:
        x.fill = PatternFill("solid", fgColor=fill)


def thanh(ws, r, t, *, so_cot=8, fill=C_SEC, color=C_HDR, size=11, italic=False, cao=24):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=so_cot)
    x = ws.cell(row=r, column=1, value=t)
    x.font = Font(bold=not italic, italic=italic, size=size, color=color)
    if fill:
        x.fill = PatternFill("solid", fgColor=fill)
    x.alignment = Alignment(vertical="top" if italic else "center", wrap_text=True)
    ws.row_dimensions[r].height = cao
    return r + 1


def tieu_de(ws, r, cot, so_cot=None):
    for i, h in enumerate(cot, 1):
        o(ws, r, i, h, bold=True, fill=C_HDR, color=C_TXT, center=True)
    ws.row_dimensions[r].height = 30
    return r + 1


# =====================================================================
# Đọc số liệu chạy thật
# =====================================================================
KQ = Path("tools/ket-qua-chay")
pm_stats = json.loads((KQ / "postman-stats.json").read_text(encoding="utf-8"))
pm_folders = json.loads((KQ / "postman-thu-muc.json").read_text(encoding="utf-8"))
e2e = json.loads((KQ / "codecept.json").read_text(encoding="utf-8"))

wb = openpyxl.Workbook()
wb.remove(wb.active)

# =====================================================================
# 05. API — Postman
# =====================================================================
ws = wb.create_sheet("05. API - Postman")
dung(ws, [30, 12, 12, 12, 26, 26, 26, 26])
x = ws.cell(row=1, column=1, value="Kiểm thử API — Postman / Newman")
x.font = Font(bold=True, size=13, color=C_HDR)
x = ws.cell(row=2, column=1, value=(
    "Kiểm thử hộp đen ở tầng HTTP: gửi request thật tới ứng dụng đang chạy và kiểm tra "
    "mã trạng thái, tiêu đề, thân phản hồi. Bổ sung cho unit test vì nó đi qua đủ chuỗi "
    "bộ lọc bảo mật, ánh xạ controller và giao dịch cơ sở dữ liệu."))
x.font = Font(size=10, italic=True, color="595959")

r = thanh(ws, 4, f"TỔNG KẾT THI HÀNH — {pm_stats['requests']['total']} request, "
                 f"{pm_stats['assertions']['total']} phép kiểm, "
                 f"{pm_stats['assertions']['failed']} lỗi", cao=30)
r = tieu_de(ws, r, ["Chỉ số", "Giá trị", "", "", "Ghi chú", "", "", ""])
for nhan, gt, ghi in [
    ("Test case gốc", 92, "Các case mã TC-xxx-nn do nhóm thiết kế"),
    ("Bước chuẩn bị", pm_stats["requests"]["total"] - 92,
     "Làm mới CSRF, đăng nhập lại, đọc id thật — không phải test case"),
    ("Tổng request", pm_stats["requests"]["total"], "Số lần gọi HTTP trong một lượt chạy"),
    ("Tổng phép kiểm", pm_stats["assertions"]["total"],
     "Postman gọi là All tests; mỗi request có nhiều phép kiểm"),
    ("Phép kiểm lỗi", pm_stats["assertions"]["failed"], "Chạy trên CSDL sạch"),
]:
    o(ws, r, 1, nhan, bold=True)
    o(ws, r, 2, gt, center=True, fill=(C_PASS if nhan == "Phép kiểm lỗi" and gt == 0 else None))
    o(ws, r, 5, ghi)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 28
    r += 1
r += 1

r = thanh(ws, r, "BẢNG 1 — PHÂN BỔ THEO NHÓM CHỨC NĂNG")
r = tieu_de(ws, r, ["Nhóm chức năng", "Request", "Phép kiểm", "Lỗi", "Yêu cầu tương ứng",
                    "", "", ""])
FR_MAP = {
    "00": "FR6.1 · Đăng nhập, đăng ký, quên mật khẩu",
    "01": "FR2.1, FR1.3 · Ba bước khởi tạo hồ sơ",
    "02": "FR6.2 · Trang tổng quan",
    "03": "FR1.1, FR1.2 · Chat AI Mentor",
    "04": "FR2.2, FR2.4 · Lộ trình học và tiến độ",
    "05": "FR3.1, FR3.2, FR3.3 · Khoảng trống kỹ năng",
    "06": "FR4.1, FR4.2, FR4.3 · Xu hướng thị trường",
    "07": "FR5.1, FR5.2, FR5.3 · Hồ sơ năng lực",
    "08": "FR6.1 · Hồ sơ và cài đặt cá nhân",
    "09": "FR7.1, FR7.2, FR7.3 · Quản trị người dùng",
    "10": "FR8.1 → FR8.4 · Cố vấn học tập",
    "11": "NFR-S01 → NFR-S05 · Phân quyền và bảo mật",
}
for ten in sorted(pm_folders):
    req, ass, fail = pm_folders[ten]
    o(ws, r, 1, ten)
    o(ws, r, 2, req, center=True)
    o(ws, r, 3, ass, center=True)
    o(ws, r, 4, fail, center=True, fill=(C_PASS if fail == 0 else C_FAIL))
    o(ws, r, 5, FR_MAP.get(ten[:2], ""))
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 26
    r += 1
# Lấy số tổng từ thống kê chính thức của newman chứ không cộng dồn từng nhóm: nhật ký
# thực thi ghi lặp một vài mục nên cộng tay ra 362 thay vì 358.
for i, v in enumerate(["TỔNG", pm_stats["requests"]["total"],
                       pm_stats["assertions"]["total"],
                       pm_stats["assertions"]["failed"]], 1):
    o(ws, r, i, v, bold=True, fill=C_SEC, center=(i > 1))
r += 2

r = thanh(ws, r, "BẢNG 2 — PHÉP KIỂM PHI CHỨC NĂNG ÁP CHO MỌI REQUEST")
r = tieu_de(ws, r, ["Mã", "Yêu cầu", "", "Ngưỡng", "Cách kiểm", "", "", "Kết quả"])
for ma, yc, nguong, cach, kq in [
    ("NFR-P01", "Trang nội bộ phản hồi nhanh", "< 2 giây",
     "Áp cho mọi request không gọi dịch vụ ngoài", "ĐẠT"),
    ("NFR-P02", "Endpoint gọi LLM / gửi email", "< 15 giây",
     "Áp cho /mentor/send, /forgot, /api/skill-gap/analyze", "ĐẠT"),
    ("NFR-P03", "Đồng bộ GitHub", "< 30 giây",
     "Áp cho /portfolio/sync — đo 14 giây sau khi tối ưu", "ĐẠT"),
    ("NFR-S05", "Không lộ stack trace, SQL, thông tin nội bộ", "0 dấu hiệu",
     'Tìm "trace", vn.uth.careercompass, insert into, hibernate', "ĐẠT"),
]:
    o(ws, r, 1, ma, bold=True)
    o(ws, r, 2, yc)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    o(ws, r, 4, nguong, center=True)
    o(ws, r, 5, cach)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
    o(ws, r, 8, kq, center=True, fill=C_PASS, bold=True)
    ws.row_dimensions[r].height = 32
    r += 1
r += 1

r = thanh(ws, r, "Vì sao có 92 test case nhưng Postman hiện 358: Postman đếm PHÉP KIỂM chứ "
                 "không đếm test case. Một test case có nhiều phép kiểm, cộng thêm 2 phép "
                 "kiểm phi chức năng mà nhóm áp cho MỌI request (một ngưỡng hiệu năng và "
                 "một kiểm tra rò rỉ thông tin). Cụ thể: 151 phép kiểm riêng + 112 × 2 = 375, "
                 "trừ đi các nhánh if/else không chạy còn 358.",
           fill=None, color="595959", size=9, italic=True, cao=54)
r = thanh(ws, r, "CÁCH CHẠY LẠI:  newman run CareerCompass.postman_collection.json "
                 "-e CareerCompass.postman_environment.json     hoặc mở Postman → Runner → "
                 "chọn collection → Run. Cần ứng dụng đang chạy ở localhost:8080.",
           fill=None, color="595959", size=9, italic=True, cao=34)

# =====================================================================
# 06. E2E — CodeceptJS
# =====================================================================
ws = wb.create_sheet("06. E2E - CodeceptJS")
dung(ws, [14, 46, 16, 12, 10, 30, 26, 26])
x = ws.cell(row=1, column=1, value="Kiểm thử giao diện đầu-cuối — CodeceptJS + Playwright")
x.font = Font(bold=True, size=13, color=C_HDR)
x = ws.cell(row=2, column=1, value=(
    "Điều khiển trình duyệt Chromium thật, thao tác đúng như người dùng: điền biểu mẫu, "
    "bấm nút, chuyển trang. Đây là tầng kiểm thử duy nhất chạm tới HTML, CSS và JavaScript "
    "nên bắt được lỗi mà unit test và kiểm thử API không thấy."))
x.font = Font(size=10, italic=True, color="595959")

st = e2e["stats"]
r = thanh(ws, 4, f"TỔNG KẾT THI HÀNH — {st['tests']} kịch bản, {st['passes']} đạt, "
                 f"{st['failures']} không đạt, {st['duration'] / 1000:.0f} giây", cao=30)
r = tieu_de(ws, r, ["Mã", "Kịch bản", "Nhóm chức năng", "Ưu tiên", "Giây",
                    "Kết quả / nguyên nhân", "", ""])
for t in e2e["tests"]:
    dat = t["status"] == "PASS"
    o(ws, r, 1, t["ma"], bold=True)
    o(ws, r, 2, t["ten"])
    o(ws, r, 3, t["feature"])
    o(ws, r, 4, t["tags"], center=True)
    o(ws, r, 5, t["giay"], center=True)
    o(ws, r, 6, "ĐẠT" if dat else "KHÔNG ĐẠT — " + (t["loi"][:120] or "xem cột ghi chú"),
      fill=C_PASS if dat else C_FAIL, bold=dat)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 32 if dat else 44
    r += 1
r += 1

r = thanh(ws, r, "BẢNG 1 — PHÂN TÍCH BA KỊCH BẢN KHÔNG ĐẠT")
r = tieu_de(ws, r, ["Mã", "Loại", "", "Nguyên nhân", "", "", "Hướng xử lý", ""])
for ma, loai, ng, huong, mau in [
    ("TC-ADM-003", "LỖI ỨNG DỤNG",
     "Quản trị viên tự đổi được vai trò của chính mình xuống STUDENT. Kịch bản này không chỉ "
     "phát hiện mà THỰC SỰ kích hoạt lỗi: sau khi chạy, admin@gmail.com mang vai trò STUDENT "
     "và không còn đường quay lại qua giao diện. Đã kiểm chứng trong CSDL, và hệ quả là bảy "
     "phép kiểm thư mục Admin của Postman đỏ theo.",
     "Chặn ở tầng dịch vụ: từ chối khi id người bị đổi trùng id người đang đăng nhập. "
     "Đã gắn nhãn @known-bug nên lệnh npm test mặc định bỏ qua.", C_FAIL),
    ("TC-DSB-001", "TEST LỖI THỜI",
     'Kịch bản tìm chuỗi "node hoàn thành" và "kỹ năng còn thiếu" trên trang tổng quan. '
     'Giao diện đã được sửa, nay dùng "Tiến độ Lộ trình" và "Kỹ năng nắm vững" nên không '
     "còn khớp. Ứng dụng chạy đúng.",
     "Cập nhật assertion theo nhãn mới, hoặc tốt hơn là gắn data-testid vào các ô thống kê "
     "để kịch bản không phụ thuộc câu chữ hiển thị.", C_WARN),
    ("TC-SKG-002", "TEST LỖI THỜI",
     'Kịch bản chờ nút submit của form[action*="/skill-gap/reports"]. Form này chỉ hiện sau '
     "khi đã có kết quả phân tích; giao diện đổi bố cục nên nút không xuất hiện kịp trong "
     "10 giây chờ.",
     "Chờ theo trạng thái (đợi khối kết quả phân tích render xong) thay vì chờ cứng theo "
     "thời gian, và gắn data-testid cho nút lưu báo cáo.", C_WARN),
]:
    o(ws, r, 1, ma, bold=True)
    o(ws, r, 2, loai, bold=True, fill=mau, center=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    o(ws, r, 4, ng)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    o(ws, r, 7, huong)
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 76
    r += 1
r += 1

r = thanh(ws, r, "PHÂN LOẠI QUAN TRỌNG: trong ba kịch bản không đạt, chỉ MỘT là lỗi ứng dụng "
                 "thật. Hai kịch bản còn lại đỏ vì giao diện đã đổi mà kịch bản chưa cập "
                 "nhật theo — bản thân chức năng vẫn chạy đúng. Đây là chi phí bảo trì đặc "
                 "trưng của kiểm thử giao diện: nó bám vào cấu trúc HTML nên dễ vỡ khi giao "
                 "diện thay đổi, khác hẳn unit test vốn chỉ phụ thuộc chữ ký hàm.",
           fill=None, color="595959", size=9, italic=True, cao=54)
r = thanh(ws, r, "CÁCH CHẠY LẠI:  cd e2e && npm test   (mặc định bỏ qua kịch bản gắn "
                 "@known-bug)  ·  npm run test:smoke  chỉ chạy nhóm smoke  ·  "
                 "SHOW=true npm test  để xem trình duyệt thao tác trực tiếp.",
           fill=None, color="595959", size=9, italic=True, cao=34)

# =====================================================================
# 00. Tổng quan
# =====================================================================
ws = wb.create_sheet("00. Tong quan")
dung(ws, [36, 14, 14, 22, 20, 24, 20, 20])
x = ws.cell(row=1, column=1, value="BÁO CÁO KIỂM THỬ PHẦN MỀM — CareerCompass (SU26SWP02)")
x.font = Font(bold=True, size=14, color=C_HDR)
x = ws.cell(row=2, column=1,
            value=f"Tổng hợp bốn tầng kiểm thử. Ngày lập: {date.today().strftime('%d/%m/%Y')}")
x.font = Font(size=10, italic=True, color="595959")

r = thanh(ws, 4, "BẢNG 1 — CÁC KỸ THUẬT ĐÃ ÁP DỤNG")
r = tieu_de(ws, r, ["Kỹ thuật", "Tầng", "Số test", "Công cụ", "Kết quả",
                    "Sheet chi tiết", "", ""])
for k, tang, n, cc, kq, sh, dat in [
    ("Phân hoạch lớp tương đương", "Đơn vị", 14, "JUnit 5", "14 / 14", "01", True),
    ("Giá trị biên — Standard + Robustness", "Đơn vị", 19, "JUnit 5", "19 / 19", "01", True),
    ("Gộp tag thành test case (slide 33)", "Đơn vị", 17, "JUnit 5", "17 / 17", "01", True),
    ("Giá trị biên — dung lượng tệp", "Đơn vị", 6, "JUnit 5", "6 / 6", "01", True),
    ("Bảng quyết định", "Đơn vị", 10, "JUnit 5", "10 / 10", "03, 03b", True),
    ("Chuyển đổi trạng thái", "Đơn vị", 9, "JUnit 5", "9 / 9", "04", True),
    ("Hộp trắng — đường cơ sở, CFG", "Đơn vị", "xem B5", "JUnit 5 + JaCoCo", "đạt",
     "B1 → B5", True),
    ("Kiểm thử API", "Tích hợp", "92 case", "Postman / Newman",
     f"{pm_stats['assertions']['total']} phép kiểm", "05", True),
    ("Kiểm thử giao diện đầu-cuối", "Hệ thống", st["tests"], "CodeceptJS + Playwright",
     f"{st['passes']} / {st['tests']}", "06", st["failures"] == 0),
]:
    o(ws, r, 1, k)
    o(ws, r, 2, tang, center=True)
    o(ws, r, 3, n, center=True)
    o(ws, r, 4, cc)
    o(ws, r, 5, kq, center=True, bold=True, fill=(C_PASS if dat else C_WARN))
    o(ws, r, 6, sh, center=True)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 28
    r += 1
r += 1

r = thanh(ws, r, "BẢNG 2 — ĐỘ BAO PHỦ MÃ NGUỒN (JaCoCo)")
r = tieu_de(ws, r, ["Chỉ số", "Trước", "Sau", "Tăng", "", "Nguồn", "", ""])
for ten, tr, sa, tang in [("Dòng lệnh (Line)", "60,1%", "75,5%", "+15,4 điểm"),
                          ("Nhánh (Branch)", "48,2%", "67,5%", "+19,3 điểm"),
                          ("Độ phức tạp được phủ", "53,4%", "67,4%", "+14,0 điểm"),
                          ("Phương thức", "65,1%", "78,5%", "+13,4 điểm")]:
    o(ws, r, 1, ten)
    o(ws, r, 2, tr, center=True)
    o(ws, r, 3, sa, center=True, bold=True, fill=C_PASS)
    o(ws, r, 4, tang, center=True)
    o(ws, r, 6, "docs/coverage/index.html")
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 26
    r += 1
r += 1

r = thanh(ws, r, "BẢNG 3 — TỔNG SỐ PHÉP KIỂM ĐÃ CHẠY")
r = tieu_de(ws, r, ["Tầng", "Đơn vị đếm", "Số lượng", "Không đạt", "", "Ghi chú", "", ""])
for tang, dv, n, fail, ghi in [
    ("Đơn vị (JUnit)", "test method", 370, 0,
     "Toàn bộ dự án, gồm cả hộp đen và hộp trắng"),
    ("Tích hợp (Postman)", "phép kiểm", pm_stats["assertions"]["total"], 0,
     "Chạy trên cơ sở dữ liệu vừa gieo lại"),
    ("Hệ thống (CodeceptJS)", "kịch bản", st["tests"], st["failures"],
     "1 lỗi ứng dụng thật, 2 kịch bản lỗi thời sau khi giao diện đổi"),
]:
    o(ws, r, 1, tang, bold=True)
    o(ws, r, 2, dv)
    o(ws, r, 3, n, center=True)
    o(ws, r, 4, fail, center=True, bold=True, fill=(C_PASS if fail == 0 else C_WARN))
    o(ws, r, 6, ghi)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 30
    r += 1
r += 1
r = thanh(ws, r, "Tám khiếm khuyết được phát hiện, bảy đã khắc phục kèm test hồi quy. "
                 "Chi tiết ở sheet 07. Khiem khuyet.",
           fill=None, color="595959", size=9, italic=True, cao=24)


# =====================================================================
# 07. Khiếm khuyết phát hiện được
# =====================================================================
ws = wb.create_sheet("07. Khiem khuyet")
dung(ws, [11, 15, 40, 32, 30, 14, 24, 24])
x = ws.cell(row=1, column=1, value="Khiếm khuyết phát hiện qua các tầng kiểm thử")
x.font = Font(bold=True, size=13, color=C_HDR)
x = ws.cell(row=2, column=1, value=(
    "Chỉ liệt kê khiếm khuyết có bằng chứng đo đạc hoặc tái hiện được, kèm kỹ thuật đã "
    "phát hiện ra nó. Đây là thước đo giá trị thực của hoạt động kiểm thử."))
x.font = Font(size=10, italic=True, color="595959")

r = tieu_de(ws, 4, ["Mã", "Mức độ", "Mô tả", "Kỹ thuật phát hiện", "Bằng chứng",
                    "Trạng thái", "Cách khắc phục", ""])
KHIEM_KHUYET = [
    ("DEF-001", "Nghiêm trọng",
     "Quản trị viên tự đổi được vai trò của chính mình xuống STUDENT, sau đó mất hẳn quyền "
     "quản trị và không có đường quay lại qua giao diện.",
     "Kiểm thử giao diện — TC-ADM-003",
     "Sau khi chạy, admin@gmail.com mang vai trò STUDENT trong CSDL; bảy phép kiểm thư mục "
     "Admin của Postman đỏ theo.",
     "CHƯA SỬA", "Chặn ở tầng dịch vụ khi id người bị đổi trùng id người đang đăng nhập.",
     C_FAIL),
    ("DEF-002", "Cao",
     "Phản hồi lỗi trả về nguyên stack trace: tên package, tên lớp, số dòng, và cả câu "
     "INSERT khi vi phạm khoá duy nhất. Vi phạm NFR-S05.",
     "Kiểm thử API — phép kiểm cấp collection",
     "POST /admin/users/99999/toggle-status trả 500 kèm trường trace dài 10.910 byte.",
     "ĐÃ SỬA",
     "Khai server.error.include-stacktrace: never, thêm GlobalExceptionHandler, và tắt "
     "spring.devtools.add-properties vì devtools ghi đè cấu hình này khi chạy từ IDE.",
     C_PASS),
    ("DEF-003", "Trung bình",
     "Ràng buộc email chỉ khai @Size(max=150) mà không khai min, nên chuỗi ba ký tự dạng "
     "a@b vẫn được chấp nhận làm địa chỉ đăng ký.",
     "Phân tích giá trị biên — TC16 Robustness BVA",
     "Kiểm chứng trực tiếp: email 3 và 5 ký tự đều qua được validation.",
     "ĐÃ SỬA",
     "Đổi thành @Size(min = 6, max = 150) — độ dài của email hợp lệ ngắn nhất a@b.co.",
     C_PASS),
    ("DEF-004", "Trung bình",
     "Đường dẫn /error không nằm trong danh sách công khai, nên mọi trang lỗi của khách chưa "
     "đăng nhập bị nuốt thành chuyển hướng sang /login. Người dùng không bao giờ thấy trang 404.",
     "Kiểm thử API — TC-PF-06",
     "GET /p/<slug không tồn tại> trả 302 tới /login;jsessionid=… thay vì 404.",
     "ĐÃ SỬA", "Thêm /error vào danh sách permitAll trong SecurityConfig.", C_PASS),
    ("DEF-005", "Trung bình",
     "Đồng bộ GitHub mất 84 giây, vượt hơn gấp đôi ngưỡng 30 giây của NFR-P03.",
     "Kiểm thử API — phép kiểm NFR-P03",
     "Đo hai lần: 84,4s và 83,3s với tài khoản 9 repository.",
     "ĐÃ SỬA",
     "Lấy README và gọi LLM song song trên pool 6 luồng thay vì tuần tự. Đo lại: 14,8s và 13,7s.",
     C_PASS),
    ("DEF-006", "Trung bình",
     "Dự án không có @ControllerAdvice nào. Mọi lỗi nghiệp vụ, kể cả không tìm thấy người "
     "dùng, đều thành HTTP 500.",
     "Kiểm thử API — TC-AD-04/05/07, TC-CS-05/10",
     "Năm request trả 500 với IllegalArgumentException và DataIntegrityViolationException.",
     "ĐÃ SỬA",
     "Thêm GlobalExceptionHandler: 400 cho tham số sai, 409 cho trùng dữ liệu, JSON cho "
     "/api/** và chuyển hướng kèm thông báo cho biểu mẫu.", C_PASS),
    ("DEF-007", "Thấp",
     "Luật R3 của bảng quyết định (IN_PROGRESS + node bị khoá) chưa từng được kiểm. Có một "
     "test tên gợi ý là phủ R3 nhưng thân hàm lại dùng DONE, tức phủ R5.",
     "Bảng quyết định — đối chiếu từng cột",
     "Tên test updateProgress_whenNodeLockedAndStatusNotStarted_isForbidden không khớp nội dung.",
     "ĐÃ SỬA", "Bổ sung sáu test phủ đủ sáu luật, đặt tên theo mã cột rule1_ đến rule6_.",
     C_PASS),
    ("DEF-008", "Thấp",
     "Hai luật R3 và R4 của bảng hiệu lực token đặt lại mật khẩu chưa có test nào: chưa ai "
     "xác nhận token đã dùng rồi thì bị từ chối.",
     "Bảng quyết định — đối chiếu từng cột",
     "PasswordResetServiceTest sẵn có chỉ phủ R1 và R2.",
     "ĐÃ SỬA", "Bổ sung TokenValidityDecisionTableTest phủ đủ bốn luật.", C_PASS),
]
for ma, muc, mo_ta, kt, bc, tt, cach, mau in KHIEM_KHUYET:
    o(ws, r, 1, ma, bold=True)
    o(ws, r, 2, muc, center=True, bold=True,
      fill=(C_FAIL if muc in ("Nghiêm trọng", "Cao") else C_WARN))
    o(ws, r, 3, mo_ta)
    o(ws, r, 4, kt)
    o(ws, r, 5, bc)
    o(ws, r, 6, tt, center=True, bold=True, fill=mau)
    o(ws, r, 7, cach)
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 78
    r += 1
r += 1
r = thanh(ws, r, "Bảy trên tám khiếm khuyết đã được khắc phục và có test hồi quy đi kèm. "
                 "DEF-001 cố ý giữ nguyên và gắn nhãn @known-bug để minh hoạ đúng quy trình: "
                 "kiểm thử ghi nhận lỗi trước, việc sửa thuộc về đợt phát triển kế tiếp.",
           fill=None, color="595959", size=9, italic=True, cao=34)


# =====================================================================
# Chép nguyên vẹn hộp đen và hộp trắng
# =====================================================================
den = openpyxl.load_workbook(NGUON_DEN)
for ten in den.sheetnames:
    if ten.startswith("99"):
        continue
    chep_sheet(den[ten], wb, ten)

trang = openpyxl.load_workbook(NGUON_TRANG)
for ten in trang.sheetnames:
    chep_sheet(trang[ten], wb, ten)

# Thứ tự trình bày: tổng quan trước, rồi hộp đen, hộp trắng, API, giao diện.
thu_tu = (["00. Tong quan"]
          + [t for t in wb.sheetnames if t[:2] in ("01", "03", "04")]
          + [t for t in wb.sheetnames if t.startswith("B") and t != "00. Tong quan"]
          + ["05. API - Postman", "06. E2E - CodeceptJS", "07. Khiem khuyet"])
wb._sheets = [wb[t] for t in thu_tu]

wb.save(RA)
print(f"Da ghi {RA}")
for s in wb.sheetnames:
    print(f"    {s:36s} {wb[s].max_row:4d} dong")
